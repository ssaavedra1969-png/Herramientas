import sys, os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import WorksheetProperties

out_path = sys.argv[1]

wb = Workbook()
ws = wb.active
ws.title = "Ingreso Vehiculos"
ws.sheet_state = "visible"
ws.freeze_panes = "A3"

headers = [
    "Patente","Marca","Modelo","Año","Chasis","Nro Motor",
    "Tipo","SubTipo","Capacidad","Carga Trompo","Kilometraje","VTV Realizacion",
    "VTV Vencimiento","VTV Costo","VTV Centro","VTV Resultado",
    "Seguro Compania","Seguro Poliza","Seguro Tipo","Seguro Vencimiento",
    "Seguro Costo","Prox Service KM","Prox Service Fecha","Conductor",
    "Empresa","Centro Trabajo","Observaciones"
]
widths = [14,18,20,8,24,22,22,16,16,16,14,18,18,14,18,16,22,20,22,18,14,16,18,22,18,18,42]
hf = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
hfill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
tw = Side(style="thin", color="FFFFFF")
hb = Border(top=tw, bottom=tw, left=tw, right=tw)
nf = Font(name="Calibri", size=10, color="999999")
na = Alignment(horizontal="center", vertical="center")

for ci, (h, w) in enumerate(zip(headers, widths), 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font, c.fill, c.alignment, c.border = hf, hfill, ha, hb
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 36

# ===== Data validations using INLINE lists =====
inline_lists = [
    ("B2:B502", "Mercedes Benz,Scania,Volvo,Iveco,Volkswagen,Ford,Chevrolet,Toyota,Fiat,Renault,Nissan,JCB,Caterpillar,Komatsu,Hyundai,New Holland,John Deere,Case,Terex,Agrale,Randon"),
    ("G2:G502", "Camion volcador,mixer,hormigonera,cisterna,jaula,playo,regador,chasis,Auto,Camioneta,Grua,Utilitario,Carga,Cargadora frontal,Retroexcavadora,Motoniveladora,Excavadora,Minicargadora,Rodillo,Acoplado,Semirremolque,Montacarga,Tolva,Achelo,Camion"),
    ("H2:H502", "Indumix,Tzr,Betonmac,tecnus,Barival,everdingm,arrastre,Montacarga,Carga,Hyva"),
    ("P2:P502", "Pendiente,Aprobado,Rechazado"),
    ("S2:S502", "Responsabilidad Civil,Todo Riesgo,Terceros Completo,Seguro Tecnico"),
    ("Z2:Z502", "Lujan,Campana,Ituzaingo,Moreno,Zarate"),
]

for rng, opts in inline_lists:
    dv = DataValidation(
        type="list",
        formula1='"' + opts + '"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valor invalido",
        error="Seleccione un valor de la lista desplegable.",
        showInputMessage=True,
        promptTitle="Lista",
        prompt="Seleccione de la lista"
    )
    dv.showDropDown = False
    ws.add_data_validation(dv)
    dv.add(rng)

# Año
dv_ano = DataValidation(
    type="whole",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Año invalido",
    error="Debe ser un numero entero (ej: 2022).",
    showInputMessage=True,
    promptTitle="Año",
    prompt="Ingrese un año (ej: 2022)"
)
dv_ano.showDropDown = False
ws.add_data_validation(dv_ano)
dv_ano.add("D2:D502")

# Capacidad
dv_cap = DataValidation(
    type="decimal",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Capacidad invalida",
    error="Debe ser un numero (ej: 25000).",
    showInputMessage=True,
    promptTitle="Capacidad",
    prompt="Ingrese la capacidad en kg"
)
dv_cap.showDropDown = False
ws.add_data_validation(dv_cap)
dv_cap.add("I2:I502")

# Kilometraje
dv_km = DataValidation(
    type="decimal",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Kilometraje invalido",
    error="Debe ser un numero (ej: 158000).",
    showInputMessage=True,
    promptTitle="Kilometraje",
    prompt="Ingrese el kilometraje"
)
dv_km.showDropDown = False
ws.add_data_validation(dv_km)
dv_km.add("K2:K502")

# VTV Costo, Seguro Costo, Prox Service KM
for col in ["N2:N502", "U2:U502", "V2:V502"]:
    dv = DataValidation(type="decimal", allow_blank=True, showErrorMessage=True, errorTitle="Valor invalido", error="Debe ser un numero.")
    dv.showDropDown = False
    ws.add_data_validation(dv)
    dv.add(col)

wb.save(out_path)
print(f"OK -> {out_path}")
